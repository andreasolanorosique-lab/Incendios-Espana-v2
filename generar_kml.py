import os
import csv
import json
import requests
import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook

from math import radians, sin, cos, sqrt, atan2, degrees, asin
MAP_KEY = os.environ["FIRMS_MAP_KEY"]
SOURCE = "VIIRS_SNPP_NRT"
BBOX = "-10,35,5,44"

DISTANCIA_AGRUPACION = 300  # metros

URL = (
    f"https://firms.modaps.eosdis.nasa.gov/api/area/csv/"
    f"{MAP_KEY}/{SOURCE}/{BBOX}/1"
)

r = requests.get(URL, timeout=120)
r.raise_for_status()

with open("fires.csv", "wb") as f:
    f.write(r.content)

kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
doc = ET.SubElement(kml, "Document")
ET.SubElement(doc, "name").text = "Incendios activos España"

BASE = "https://andreasolanorosique-lab.github.io/Incendios-Espana-v2/icons/"

styles = {
    "pink": ("fire_pink.png", 1.0),
    "green": ("fire_green.png", 1.0),
    "yellow": ("fire_yellow.png", 1.0),
    "orange": ("fire_orange.png", 1.0),
    "red": ("fire_red.png", 1.0),
}

for sid, (icon, scale) in styles.items():
    st = ET.SubElement(doc, "Style", id=sid)

    iconstyle = ET.SubElement(st, "IconStyle")
    ET.SubElement(iconstyle, "scale").text = str(scale)

    icono = ET.SubElement(iconstyle, "Icon")
    ET.SubElement(icono, "href").text = BASE + icon

    label = ET.SubElement(st, "LabelStyle")
    ET.SubElement(label, "scale").text = "0"


# =====================================================
# FUNCIONES DE AGRUPACIÓN
# =====================================================

def distancia_metros(lat1, lon1, lat2, lon2):

    R = 6371000

    lat1 = radians(lat1)
    lon1 = radians(lon1)
    lat2 = radians(lat2)
    lon2 = radians(lon2)

    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = (
        sin(dlat / 2) ** 2
        + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    )

    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    return R * c
def crear_circulo(lat, lon, radio_m, pasos=36):

    R = 6371000

    puntos = []

    lat1 = radians(lat)
    lon1 = radians(lon)

    for i in range(pasos + 1):

        ang = radians(i * 360 / pasos)

        lat2 = asin(
            sin(lat1) * cos(radio_m / R)
            + cos(lat1) * sin(radio_m / R) * cos(ang)
        )

        lon2 = lon1 + atan2(
            sin(ang) * sin(radio_m / R) * cos(lat1),
            cos(radio_m / R) - sin(lat1) * sin(lat2)
        )

        puntos.append((degrees(lon2), degrees(lat2)))

    return puntos
def agrupar_focos(focos):

    grupos = []

    for foco in focos:

        añadido = False

        for grupo in grupos:

            for referencia in grupo:

                if distancia_metros(
                    foco["lat"],
                    foco["lon"],
                    referencia["lat"],
                    referencia["lon"]
                ) <= DISTANCIA_AGRUPACION:

                    grupo.append(foco)
                    añadido = True
                    break

            if añadido:
                break

        if not añadido:
            grupos.append([foco])

    return grupos
def cargar_municipios():
    wb = load_workbook("IGN_INFOGEO_MUNICIPIOS.xlsx", read_only=True, data_only=True)
    ws = wb.active

    municipios = []

    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    idx_nombre = encabezados.index("Nombre")
    idx_mapa = encabezados.index("Ver en mapa")

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre = fila[idx_nombre]
        url = fila[idx_mapa]

        if not nombre or not url:
            continue

        m = re.search(r"center=([-\d\.]+),([-\d\.]+)", url)
        if not m:
            continue

        lon = float(m.group(1))
        lat = float(m.group(2))

        municipios.append({
            "nombre": nombre,
            "lat": lat,
            "lon": lon
        })

    wb.close()
    return municipios

# =====================================================
# LEER TODOS LOS FOCOS NASA
# =====================================================

focos = []

with open("fires.csv", encoding="utf-8") as f:

    for row in csv.DictReader(f):

        lat = row.get("latitude")
        lon = row.get("longitude")

        if not lat or not lon:
            continue

        try:
            frp = float(row.get("frp") or 0)
        except:
            frp = 0

        focos.append({
            "lat": float(lat),
            "lon": float(lon),
            "frp": frp,
            "row": row
        })


# =====================================================
# AGRUPAR FOCOS
# =====================================================
# =====================================================
# CARGAR MUNICIPIOS
# =====================================================

municipios = cargar_municipios()
print(f"Municipios cargados: {len(municipios)}")
grupos = agrupar_focos(focos)

grupos = [
    grupo
    for grupo in grupos
    if len(grupo) > 1 or max(f["frp"] for f in grupo) >= 50
]

# =====================================================
# CREAR PLACEMARKS
# (EN ESTA FASE SIGUE SIENDO UNO POR FOCO)
# =====================================================

for grupo in grupos:

    cantidad = len(grupo)

    lat = sum(f["lat"] for f in grupo) / cantidad
    lon = sum(f["lon"] for f in grupo) / cantidad

    foco_principal = max(grupo, key=lambda f: f["frp"])

    row = foco_principal["row"]
    frp = foco_principal["frp"]

    if frp < 50:
        style = "#pink"
        confianza = "Baja"

    elif frp < 100:
        style = "#yellow"
        confianza = "Media"

    elif frp < 200:
        style = "#orange"
        confianza = "Alta"

    else:
        style = "#red"
        confianza = "Muy alta"

    pm = ET.SubElement(doc, "Placemark")
    ET.SubElement(pm, "styleUrl").text = style
    ET.SubElement(pm, "name").text = ""

    desc = f"""
    <![CDATA[
    <h2>🔥 Incendio activo</h2>
    <table border="0" cellpadding="4">
    <tr><td><b>FRP</b></td><td>{frp:.1f} MW</td></tr>
    <tr><td><b>Confianza</b></td><td>{confianza}</td></tr>
    <tr><td><b>Fecha</b></td><td>{row.get('acq_date','')}</td></tr>
    <tr><td><b>Hora</b></td><td>{row.get('acq_time','')} UTC</td></tr>
    <tr><td><b>Satélite</b></td><td>{row.get('satellite','')}</td></tr>
    <tr><td><b>Instrumento</b></td><td>{row.get('instrument','')}</td></tr>
    <tr><td><b>Latitud</b></td><td>{lat}</td></tr>
    <tr><td><b>Longitud</b></td><td>{lon}</td></tr>
    </table>
    ]]>
    """

    ET.SubElement(pm, "description").text = desc

    pt = ET.SubElement(pm, "Point")
    ET.SubElement(pt, "coordinates").text = f"{lon},{lat},0"
    # ===== CÍRCULO DE PRUEBA =====

    if grupo == grupos[0]:

        circulo = crear_circulo(lat, lon, 1000)

        pol = ET.SubElement(doc, "Placemark")

        estilo = ET.SubElement(pol, "Style")

        linea = ET.SubElement(estilo, "LineStyle")
        ET.SubElement(linea, "color").text = "ff0000ff"
        ET.SubElement(linea, "width").text = "2"

        relleno = ET.SubElement(estilo, "PolyStyle")
        ET.SubElement(relleno, "color").text = "300000ff"

        polygon = ET.SubElement(pol, "Polygon")

        outer = ET.SubElement(polygon, "outerBoundaryIs")
        ring = ET.SubElement(outer, "LinearRing")

        ET.SubElement(ring, "coordinates").text = "\n".join(
            f"{lon},{lat},0" for lon, lat in circulo
        )
# =====================================================
# CARGAR RED DE GASODUCTOS
# =====================================================

with open(
    "infraestructuras/gasoductos/gasoductos.json",
    encoding="utf-8"
) as f:
    gasoductos = json.load(f)

# =====================================================
# DIBUJAR RED DE GASODUCTOS
# =====================================================

carpeta_gas = ET.SubElement(doc, "Folder")
ET.SubElement(carpeta_gas, "name").text = "Gasoductos"

for tramo in gasoductos:

    coords = tramo.get("coordenadas", [])

    if len(coords) < 2:
        continue

    pm = ET.SubElement(carpeta_gas, "Placemark")
    ET.SubElement(pm, "name").text = tramo.get("nombre", "Gasoducto")

    estilo = ET.SubElement(pm, "Style")
    linea = ET.SubElement(estilo, "LineStyle")

    ET.SubElement(linea, "color").text = "ff0000ff"
    ET.SubElement(linea, "width").text = "4"

    ls = ET.SubElement(pm, "LineString")

    ET.SubElement(ls, "tessellate").text = "1"
    ET.SubElement(ls, "altitudeMode").text = "clampToGround"

    ET.SubElement(ls, "coordinates").text = "\n".join(
        f"{lon},{lat}"
        for lon, lat in coords
    )

tree = ET.ElementTree(kml)

try:
    ET.indent(tree, space="  ")
except AttributeError:
    pass
tree.write(
    "incendios_actual.kml",
    encoding="utf-8",
    xml_declaration=True
)
